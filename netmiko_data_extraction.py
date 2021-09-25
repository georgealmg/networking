#!/usr/bin/env python3
#v1.1.2

import os, getpass, sys, re
from datetime import datetime
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoTimeoutException, SSHException, AuthenticationException
from openpyxl import Workbook

windowsuser = os.getlogin()
os.chdir("C:/Users/"+windowsuser+"/Documents")
user = input("Username: ")
pas = getpass.getpass()
sw = []
sw_gw = ""
list_of_int = []
list_of_mac = []
list_of_ip = []
list_of_vlan = []
list_of_speeds = []
list_of_duplex = []
list_of_modes = []
list_of_description = []
list_of_states = []
list_of_host = []
list_of_sw = []
list_of_channel = []
list_of_sfp = []
list_of_running = []
total_sw = len(sw)
tiempo1 = datetime.now()
tiempo_inicial = tiempo1.strftime("%H:%M:%S")
print(f"La ejecucion de este programa inicio a las {tiempo_inicial}, se validara un total de {str(total_sw)} switch.")

def data_extraction(conn,inter,list_of_uplinks,hostname):
    #Condicion aplica para equipos Nexus debido a sus output distintos.
    if sw_i in []:
        #Validacion de estado de interfaz.
        state = conn.send_command(f"show interface {inter} | include down")
        if "authorization failed" in state:
            list_of_states.append("Error de autorización")
        elif "is down" in state:
            state_0 = re.search(r"\(\D+\)",state)
            if state_0 == None:
                list_of_states.append("(down)")
            else:
                state = state_0.group()
                list_of_states.append(state)
        else:
            list_of_states.append("(connected)")
        #Validacion de configuracion de port-channel.
        if "Po" not in inter:
            channel = conn.send_command(f"show running-config interface {inter} | include channel")
            if "authorization failed" in channel:
                list_of_channel.append("Error de autorización")
            else:
                port_0 = re.search(r"\d+",channel)
                if port_0 != None:
                    port = port_0.group()
                    if "active" in channel or "passive" in channel:
                        channel_type = "LACP"
                    elif "auto" in channel or "desirable" in channel:
                        channel_type = "PAgP"
                    elif "on" in channel:
                        channel_type = "on"
                    channel = "Po"+ str(port) + " "+ channel_type
                    list_of_channel.append(channel)
                elif port_0 == None:
                    list_of_channel.append("Interfaz standalone")
        elif "Po" in inter:
            channel = conn.send_command(f"show port-channel summary | include {inter}")
            if "authorization failed" in channel:
                list_of_channel.append("Error de autorización")
            else:
                port_0 = re.search(r"Po\d+\(\w+\)",channel)
                port = port_0.group()
                if "LACP" in channel:
                    channel_type = "LACP"
                elif "PAgP" in channel:
                    channel_type = "PAgP"
                else:
                    channel_type = "on"
                channel = port + " "+ channel_type
            list_of_channel.append(channel)
    #Condicion aplica para equipos IOS.
    else:
        #Validacion de estado de interfaz.
        state = conn.send_command(f"show interface {inter} | include line")
        if "authorization failed" in state:
            list_of_states.state("Error de autorización")
        elif "is up" in state:
            list_of_states.append("(connected)")
        else:
            state_0 = re.search(r"\(\D+\)",state)
            if state_0 == None:
                list_of_states.append("(down)")
            else:
                state = state_0.group()
                list_of_states.append(state)
        #Validacion de configuracion de port-channel.
        if "Po" not in inter:
            channel = conn.send_command(f"show running-config interface {inter} | include channel")
            if "authorization failed" in channel:
                list_of_channel.append("Error de autorización")
            else:
                port_0 = re.search(r"\d+",channel)
                if port_0 != None:
                    port = port_0.group()
                    if "active" in channel or "passive" in channel:
                        channel_type = "LACP"
                    elif "auto" in channel or "desirable" in channel:
                        channel_type = "PAgP"
                    elif "on" in channel:
                        channel_type = "on"
                    channel = "Po"+ str(port) + " "+ channel_type
                    list_of_channel.append(channel)
                elif port_0 == None:
                    list_of_channel.append("Interfaz standalone")
        elif "Po" in inter:
            channel = conn.send_command(f"show etherchannel summary | include {inter}")
            if "authorization failed" in channel:
                list_of_channel.append("Error de autorización")
            else:
                port_0 = re.search(r"Po\d+\(\w+\)",channel)
                port = port_0.group()
                if "LACP" in channel:
                    channel_type = "LACP"
                elif "PAgP" in channel:
                    channel_type = "PAgP"
                else:
                    channel_type = "on"
                channel = port + " "+ channel_type
                list_of_channel.append(channel)
    #Extraccion de hostname.
    list_of_host.append(hostname.strip("#"))
    list_of_sw.append(sw_i)
    if inter in list_of_uplinks:
        list_of_mac.append("Interfaz de interconexion")
    #Validacion de MAC en interfaces de acceso.
    elif inter not in list_of_uplinks:
        mac_add = conn.send_command(f"show mac address interface {inter}")
        if "authorization failed" in mac_add:
            list_of_mac.append("Error de autorización")
        else:
            mac_int = re.findall(r"\w+[.]\w+[.]\w*", mac_add)
            if mac_int == []:
                list_of_mac.append("Interfaz sin MAC")
            elif mac_int != []:
                list_of_mac.append(mac_int)
    list_of_int.append(inter)
    #Validacion de medio fisico.
    status = conn.send_command(f"show interface {inter} status")
    if "authorization failed" in status:
        list_of_sfp.append("Error de autorización")
    elif "aseT" in status or "-T" in status:
        list_of_sfp.append("TX")
    elif "ase-SR" in status:
        list_of_sfp.append("SR")
    elif "ase-SX" in status or "BaseSX" in status:
        list_of_sfp.append("SX")
    elif "Po" in inter:
        list_of_sfp.append("Port-channel")
    else:
        list_of_sfp.append("Not Present")
    #Validacion de estado de switchport.
    if "authorization failed" in status:
        list_of_modes.append("Error de autorización")
        list_of_vlan.append("Error de autorización")
    elif "trunk" in status and "authorization failed" not in status:
        list_of_modes.append("trunk")
        #Extraccion de VLAN.
        running_config = conn.send_command(f"show running-config interface {inter} | in allowed")
        if "authorization failed" in running_config:
            list_of_vlan.append("Error de autorización")
        else:
            vlan = re.findall(r"\d+", running_config)
            if vlan != []:
                list_of_vlan.append(vlan)
            elif vlan == []:
                list_of_vlan.append("Interfaz sin VLAN")
    elif "trunk" not in status and "authorization failed" not in status:
        list_of_modes.append("access")
        #Extraccion de VLAN.
        running_config = conn.send_command(f"show running-config interface {inter} | in access")
        if "authorization failed" in running_config:
            list_of_vlan.append("Error de autorización")
        vlan = re.findall(r"\d+", running_config)
        if vlan != []:
            list_of_vlan.append(vlan)
        elif vlan == []:
            list_of_vlan.append("Interfaz sin VLAN")
    #Validacion de duplex.
    if "authorization failed" in status:
        list_of_duplex.append("Error de autorización")
    elif "full" in status:
        list_of_duplex.append("Full")
    elif "half" in status:
        list_of_duplex.append("Half")
    elif "auto" in status:
        list_of_duplex.append("Auto")
    #Validacion de velocidades.
    speed_str = conn.send_command(f"show interface {inter} | include BW")
    if "authorization failed" in speed_str:
        list_of_speeds.append("Error de autorización")
    else:
        bw_0 = re.search(r"BW \d+", speed_str)
        bw_1 = bw_0.group()
        speed = int(bw_1.strip("BW "))//1000
        list_of_speeds.append(speed)
    #Extraccion de descripcion.
    description_0 = conn.send_command(f"show interface {inter} | in Description")
    if "authorization failed" in description_0:
        list_of_description.append("Error de autorización")
    elif description_0 != "":
        description = description_0.strip("Description: ")
        list_of_description.append(description)
    elif description_0 == "":
        list_of_description.append("Sin descripcion")
    #Extraccion de configuracion de interfaz.    
    running_config = conn.send_command(f"show running-config interface {inter}")
    if "authorization failed" in running_config:
        list_of_running.append("Error de autorización")
    else:
        list_of_running.append(running_config)

def validacion(conn):
    hostname = conn.find_prompt()
    print(f"Ejecutando en {hostname} : {sw_i}")
    if sw_i == "":
        list_of_uplinks = ["Gix/x"]
    status = conn.send_command("show interface status")
    if sw_i == "":
        gi2_interfaces = re.findall(r"Gi\d/\d/\d+", status)
        for inter in gi2_interfaces:
            data_extraction(conn,inter,list_of_uplinks,hostname)
        portchannel_interfaces = re.findall(r"Po\d+", status)
        for inter in portchannel_interfaces:
            data_extraction(conn,inter,list_of_uplinks,hostname)
    else:
        te_interfaces = re.findall(r"Te\d/\d+", status)
        for inter in te_interfaces:
            data_extraction(conn,inter,list_of_uplinks,hostname)
        gi_interfaces = re.findall(r"Gi\d/\d+", status)
        for inter in gi_interfaces:
            data_extraction(conn,inter,list_of_uplinks,hostname)
        fa_interfaces = re.findall(r"Fa\d/\d+", status)
        for inter in fa_interfaces:
            data_extraction(conn,inter,list_of_uplinks,hostname)
        eth_interfaces = re.findall(r"Eth\d/\d+", status)
        for inter in eth_interfaces:
            data_extraction(conn,inter,list_of_uplinks,hostname)
        portchannel_interfaces = re.findall(r"Po\d+", status)
        for inter in portchannel_interfaces:
            data_extraction(conn,inter,list_of_uplinks,hostname)
    print(f"Ejecucion finalizada en {hostname} : {sw_i}")
    conn.disconnect()

for sw_i in sw:
    try:
        conn = ConnectHandler(device_type= "cisco_ios_ssh",host= sw_i,username= user,password= pas, fast_cli=False)
        validacion(conn)
    except(AuthenticationException):
        sys.exit("Por favor ejecute nuevamente el programa ya que introdujo una contraseña erronea.")
    except(SSHException, NetMikoTimeoutException):
        try:
            conn = ConnectHandler(device_type="cisco_ios_telnet", host= sw_i, username= user, password= pas, fast_cli=False)
            validacion(conn)
        except:
            print(f"Switch con IP: {sw_i} esta fuera de linea")

#Extraccion de direcciones IP.
def direccion_IP(conn):
    hostname_gw = conn.find_prompt()
    print(f"Validando tabla de ARP en {hostname_gw} : {sw_gw}")
    for mac in list_of_mac:
        arp_list = []
        if mac == "Interfaz sin MAC":
            list_of_ip.append("Sin IP")
        elif mac == "Interfaz de interconexion":
            list_of_ip.append("Interfaz de interconexion")
        else:
            for m in mac:
                if m == "ffff.ffff.ffff":
                    arp_list.append("MAC Broadcast")
                else:
                    arp = conn.send_command(f"show ip arp | in {m}")
                    if "authorization failed" in arp:
                        arp_list.append("Error de autorización")
                    else:
                        ip = re.findall(r"\d+[.]\d+[.]\d+[.]\d+", arp)
                        if ip != []:
                            arp_list.append(ip)
                        elif ip == []:
                            arp_list.append("Sin IP")
            list_of_ip.append(arp_list)
    print(f"Validacion de ARP finalizada en {hostname_gw} : {sw_gw}")
    conn.disconnect()

try:
    conn = ConnectHandler(device_type= "cisco_ios_ssh",host= sw_gw,username= user,password= pas, fast_cli=False)
    direccion_IP(conn)
except(AuthenticationException):
    sys.exit("Por favor ejecute nuevamente el programa ya que introdujo una contraseña erronea.")
except(SSHException, NetMikoTimeoutException):
    try:
        conn = ConnectHandler(device_type="cisco_ios_telnet", host= sw_gw, username= user, password= pas, fast_cli=False)
        direccion_IP(conn)
    except:
        print(f"Switch con IP: {sw_gw} esta fuera de linea")

wb = Workbook()
dest_filename = 'Levantamiento.xlsx'
ws1 = wb.active
ws1.title = "Data"
_= ws1.cell(column=1, row=1, value= "Switch")
_= ws1.cell(column=2, row=1, value= "IP del switch")
_= ws1.cell(column=3, row=1, value= "IP del servidor")
_= ws1.cell(column=4, row=1, value= "MAC del servidor")
_= ws1.cell(column=5, row=1, value= "Interfaz de SW")
_= ws1.cell(column=6, row=1, value= "VLAN")
_= ws1.cell(column=7, row=1, value= "Velocidad (Mbps)")
_= ws1.cell(column=8, row=1, value= "Duplex")
_= ws1.cell(column=9, row=1, value= "Access o Trunk")
_= ws1.cell(column=10, row=1, value= "Description")
_= ws1.cell(column=11, row=1, value= "Estado de interfaz")
_= ws1.cell(column=12, row=1, value= "Port-channel")
_= ws1.cell(column=13, row=1, value= "SFP")
_= ws1.cell(column=14, row=1, value= "Configuracion")
row_value = 2
iterator = 0
while iterator <= (len(list_of_int) - 1):
    try:
        _= ws1.cell(column=1, row=row_value, value=list_of_host[iterator])
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=2, row=row_value, value=list_of_sw[iterator])
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=3, row=row_value, value=str(list_of_ip[iterator]))
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=4, row=row_value, value=str(list_of_mac[iterator]))
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=5, row=row_value, value=list_of_int[iterator])
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=6, row=row_value, value=str(list_of_vlan[iterator]))
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=7, row=row_value, value=list_of_speeds[iterator])
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=8, row=row_value, value=list_of_duplex[iterator])
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=9, row=row_value, value=list_of_modes[iterator])
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=10, row=row_value, value=str(list_of_description[iterator]))
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=11, row=row_value, value=str(list_of_states[iterator]))
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=12, row=row_value, value=list_of_channel[iterator])
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=13, row=row_value, value=list_of_sfp[iterator])
    except(IndexError):
        pass
    try:
        _= ws1.cell(column=14, row=row_value, value=list_of_running[iterator])
    except(IndexError):
        pass
    iterator+=1
    row_value+=1
wb.save(filename = dest_filename)
tiempo2 = datetime.now()
tiempo_final = tiempo2.strftime("%H:%M:%S")
print(f"La ejecucion de este programa finalizo a las {tiempo_final}, se validaron un total de {str(total_sw)} equipo")
tiempo_ejecucion = tiempo2 - tiempo1
print(f"El tiempo de ejecucion del programa fue de: {tiempo_ejecucion}")
