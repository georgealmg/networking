#!/usr/bin/env python3
#v1.0.1

import os, concurrent.futures, sys
from openpyxl import load_workbook
from getpass import getpass
from datetime import datetime
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoTimeoutException, SSHException, AuthenticationException

def DC(sw,user,pas,sw_out):
    def respaldos_DC(conn):
        hostname = (conn.find_prompt())
        respaldos_txt = open(hostname.strip("#")+".txt","+w")
        respaldos_txt.write(conn.send_command("show running-config"))
        respaldos_txt.close()
        try:
            conn.save_config()
        except(OSError):
            conn.save_config(confirm=True,confirm_response="yes\r")
        conn.disconnect()
        print(f"Configuracion de {hostname} respaldada.")
    try:
        conn = ConnectHandler(device_type= "cisco_ios_ssh",host= sw,username= user,password= pas)
        respaldos_DC(conn)
    except(AuthenticationException):
        sys.exit("Por favor ejecute nuevamente el programa ya que introdujo una contraseña erronea.")
    except(NetMikoTimeoutException,SSHException):
        try:
            conn = ConnectHandler(device_type= "cisco_ios_telnet",host= sw,username= user,password= pas)
            respaldos_DC(conn)
        except:
            print(f"Switch con IP: {str(sw)} esta fuera de linea.")
            sw_out.append(sw)

def main():
    windowsuser = os.getlogin()
    try:
        os.chdir("C:/Users/"+str(windowsuser)+"/Documents")
    except(FileNotFoundError):
        os.chdir(os.getcwd())
    user = input("Username: ")
    pas = getpass()
    sw_dc = []
    sw_out = []
    fp = open("sw_out.txt","+w")
    excel_file = load_workbook("Respaldos.xlsx")
    sheet = excel_file["IP"]
    for ip in range(2, 999999):
        valor = sheet.cell(row=ip, column=1).value
        if valor != None and valor not in sw_dc:
            sw_dc.append(valor)
        elif valor == None:
            break
    tiempo1 = datetime.now()
    tiempo_inicial = tiempo1.strftime("%H:%M:%S")
    print(f"La ejecucion de este programa inicio a las {tiempo_inicial}, se validara un total de {len(sw_dc)} switch.")
    
    fecha = datetime.now()
    directorio_fecha = fecha.strftime("%d-%m-%y")
    directorio = "Data_Center"
    try:
        os.mkdir(f"C:/Users/{windowsuser}/Documents/{directorio}")
    except(FileExistsError):
        print(f"La carpeta {directorio} ya existe.")
    try:
        os.mkdir(f"C:/Users/{windowsuser}/Documents/{directorio}/{directorio_fecha}")
    except(FileExistsError):
        print(f"La carpeta {directorio}/{directorio_fecha} ya existe.")
    os.chdir(f"C:/Users/{windowsuser}/Documents/{directorio}/{directorio_fecha}")
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor_nx:
        ejecucion_nx = {executor_nx.submit(DC,sw,user,pas,sw_out,directorio,directorio_fecha): sw for sw in sw_dc}
    for output_nx in concurrent.futures.as_completed(ejecucion_nx):
        output_nx.result()

    print(f"Equipos caidos: {sw_out}")
    fp.write("Equipos caidos:"+"\n")
    fp.close()
    tiempo2 = datetime.now()
    tiempo_final = tiempo2.strftime("%H:%M:%S")
    print(f"La ejecucion de este programa finalizo a las {tiempo_final}, se valido un total de {len(sw_dc)} switch de los cuales {len(sw_out)} estan fuera de linea.")
    tiempo_ejecucion = tiempo2 - tiempo1
    print(f"El tiempo de ejecucion del programa fue de: {tiempo_ejecucion}")

if __name__ == "__main__":
    main()