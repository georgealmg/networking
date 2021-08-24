#!/usr/bin/env python3
#v1.0.2

import os, sys, re, concurrent.futures
from getpass import getpass
from openpyxl import  load_workbook
from datetime import datetime
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoTimeoutException, SSHException, AuthenticationException

def NXOS(sw_n,user,pas,sw_out,txt_file):
    def puertas_NXOS(conn):
        hostname = (conn.find_prompt())
        down = conn.send_command("show interface status | ex connected | ex  Po | ex mgmt | ex Vl")
        interfaces = re.findall(r"Fa\d/\d+/?\d*|Gi\d/\d+/?\d*|Te\d/\d+/?\d*|Eth\d+/\d+/?\d*",down)
        for inter in interfaces:
            last = conn.send_command(f"show interface {inter} | include Last")
            if "Last link flapped never" in last:
                inter_descrip = conn.send_command(f"show interface {inter} | include Description")
                description = re.search("(Description:) (.+)",inter_descrip)
                txt_file.write(f"{hostname} {description.group(2)}\n")
        print(f"Ejecucion finalizada en {hostname}.")
        conn.disconnect()
    try:
        conn = ConnectHandler(device_type= "cisco_nxos_ssh",host= sw_n,username= user,password= pas)
        puertas_NXOS(conn)
    except(AuthenticationException):
        sys.exit("Por favor ejecute nuevamente el programa ya que introdujo una contraseña erronea.")
    except(NetMikoTimeoutException,SSHException):
        print(f"Switch con IP: {str(sw_n)} esta fuera de linea.")
        sw_out.append(sw_n)
        
def IOS(sw_i,user,pas,sw_out,txt_file):
    def puertas_IOS(conn):
        hostname = (conn.find_prompt())
        down = conn.send_command("show interfaces status | exclude connected|Po|")
        interfaces = re.findall(r"Fa\d/\d+/?\d*|Gi\d/\d+/?\d*|Te\d/\d+/?\d*|Eth\d+/\d+/?\d*",down)
        for inter in interfaces:
            last = conn.send_command(f"show interface {inter} | include Last")
            if "Last link flapped never" in last:
                inter_descrip = conn.send_command(f"show interface {inter} | include Description")
                description = re.search("(Description:) (.+)",inter_descrip)
                txt_file.write(f"{hostname} {description.group(2)}\n")
        print(f"Ejecucion finalizada en {hostname}.")
        conn.disconnect()
    try:
        conn = ConnectHandler(device_type= "cisco_ios_ssh",host= sw_i,username= user,password= pas)
        puertas_IOS(conn)
    except(AuthenticationException):
        print(f"Error de autenticacion en IP: {sw_i}.")
    except(NetMikoTimeoutException,SSHException):
        try:
            conn = ConnectHandler(device_type= "cisco_ios_telnet",host= sw_i,username= user,password= pas)
            puertas_IOS(conn)
        except:
            print(f"Switch con IP: {str(sw_i)} esta fuera de linea.")
            sw_out.append(sw_i)

def main():
    windowsuser = os.getlogin()
    try:
        os.chdir("C:/Users/"+str(windowsuser)+"/Documents")
    except(FileNotFoundError):
        os.chdir(os.getcwd())
    user = input("Username: ")
    pas = getpass()
    sw_nx = []
    sw_ios = []
    sw_out = []
    txt_file = open("puertas-disponibles.txt", "w+")
    excel_file = load_workbook("switch.xlsx")
    sheet = excel_file["Puertas"]
    for x in range(2, 999999):
        valor = sheet.cell(row=x, column=2).value
        if valor != None and valor not in sw_ios:
            sw_ios.append(valor)
        elif valor == None:
            break
    for x in range(2, 999999):
        valor = sheet.cell(row=x, column=3).value
        if valor != None and valor not in sw_nx:
            sw_nx.append(valor)
        elif valor == None:
            break
    total_sw = len(sw_nx) + len(sw_ios)
    tiempo1 = datetime.now()
    tiempo_inicial = tiempo1.strftime("%H:%M:%S")
    print(f"La ejecucion de este programa inicio a las {tiempo_inicial}, se validara un total de {str(total_sw)} switch.")

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor_nx:
        ejecucion_nx = {executor_nx.submit(NXOS,sw_n,user,pas,sw_out,txt_file): sw_n for sw_n in sw_nx}
    for output_nx in concurrent.futures.as_completed(ejecucion_nx):
        output_nx.result()
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor_ios:
        ejecucion_ios = {executor_ios.submit(IOS,sw_i,user,pas,sw_out,txt_file): sw_i for sw_i in sw_ios}
    for output_ios in concurrent.futures.as_completed(ejecucion_ios):
        output_ios.result()

    print(f"Equipos caidos: {sw_out}")
    txt_file.write("Equipos caidos:\n"+str(sw_out))
    txt_file.close()
    tiempo2 = datetime.now()
    tiempo_final = tiempo2.strftime("%H:%M:%S")
    print(f"La ejecucion de este programa finalizo a las {tiempo_final}, se valido un total de {str(total_sw)} switch de los cuales {str(len(sw_out))} estan fuera de linea.")
    tiempo_ejecucion = tiempo2 - tiempo1
    print(f"El tiempo de ejecucion del programa fue de: {tiempo_ejecucion}")

if __name__ == "__main__":
    main()
