#!/usr/bin/env python3
#v1.0.8

import csv,concurrent.futures, os, re, sys
from getpass import getpass, getuser
from datetime import datetime
from netmiko import ConnectHandler
from netmiko.ssh_exception import SSHException, AuthenticationException
from pandas import read_csv
from openpyxl import load_workbook


def inter_output(conn,description,hostname,sw,poe,writer):
    if description == "":
        return
    elif description != "":
        port = re.search(r"(Fa\d/\d+/?\d*|Gi\d/\d+/?\d*)(\s+)(admin down|down)(\s+)(down)(\s+)(.+)",description)
    if port.group(1) in poe:
        poe = "PoE"
    elif port.group(1) not in poe:
        poe = "no PoE"
    vlan = re.search(r"(switchport)(\s+)(access)(\s+)(vlan)(\s+)(\d+)",conn.send_command(f"show run inter {port.group(1)}"))
    try:
        vlan = vlan.group(7)
    except(AttributeError):
        vlan = "1"
    writer.writerow({"Hostname":hostname,"Hostaddress":sw,"Port":port.group(1),"PoE":poe,"Status":port.group(3),"Description":port.group(7),"VLAN":vlan})
        
def suc(conn,sw,writer):
    hostname = conn.find_prompt()
    print(f"Validacion iniciada --> {hostname}.")
    poe = conn.send_command("show power inline")
    down = conn.send_command("show interface description | include admin")
    interfaces = re.findall(r"Fa\d/\d+/?\d*|Gi\d/\d+/?\d*",down)
    for inter in interfaces:
        description = conn.send_command(f"show interface {inter} description | in /")
        inter_output(conn,description,hostname,sw,poe,writer)
    notconnect = conn.send_command("show inter status | in notconnect")
    interfaces = re.findall(r"Fa\d/\d+/?\d*|Gi\d/\d+/?\d*",notconnect)
    for inter in interfaces:
        last = conn.send_command(f"show interface {inter} | in Last")
        if "y" in last or "Last input never, output never" in last:
            description = conn.send_command(f"show interface {inter} description | in /")
            inter_output(conn,description,hostname,sw,poe,writer)
    print(f"Validacion finalizada --> {hostname}.")
    conn.disconnect()

def connection(sw,user,pas,sw_out,swout_file,writer):
    try:
        conn = ConnectHandler(device_type= "cisco_ios_ssh",host= sw,username= user,password= pas, fast_cli= False)
        suc(conn,sw,writer)
    except(ConnectionRefusedError):
        sw_out.append(sw)
        print(f"Error:{sw}:ConnectionRefused error")
        swout_file = open("sw_out.txt","+a")
        swout_file.write(f"Error:{sw}:ConnectionRefused error"+"\n")
        swout_file.close()
    except(AuthenticationException):
        sw_out.append(sw)
        print(f"Error:{sw}:Authentication error")
        swout_file = open("sw_out.txt","+a")
        swout_file.write(f"Error:{sw}:Authentication error"+"\n")
        swout_file.close()
    except(SSHException):
        try:
            conn = ConnectHandler(device_type= "cisco_ios_telnet",host= sw,username= user,password= pas,fast_cli= False)
            suc(conn,sw,writer)
        except(ConnectionRefusedError):
            sw_out.append(sw)
            print(f"Error:{sw}:ConnectionRefused error")
            swout_file = open("sw_out.txt","+a")
            swout_file.write(f"Error:{sw}:ConnectionRefused error"+"\n")
            swout_file.close()
        except(TimeoutError):
            sw_out.append(sw)
            print(f"Error:{sw}:Timeout error")
            swout_file = open("sw_out.txt","+a")
            swout_file.write(f"Error:{sw}:Timeout error"+"\n")
            swout_file.close()
        except(AuthenticationException):
            sw_out.append(sw)
            print(f"Error:{sw}:Authentication error")
            swout_file = open("sw_out.txt","+a")
            swout_file.write(f"Error:{sw}:Authentication error"+"\n")
            swout_file.close()
    except(EOFError):
        sw_out.append(sw)
        print(f"Error:{sw}:EOF error")
        swout_file = open("sw_out.txt","+a")
        swout_file.write(f"Error:{sw}:EOF error"+"\n")
        swout_file.close()

def main():
    try:
       os.chdir(f"/mnt/c/Users/{getuser()}/Documents/networking")
    except(FileNotFoundError):
        os.chdir(os.getcwd())
    user = input("Username: ")
    pas = getpass()
    sw_ios = []
    sw_out = []
    swout_file = open("sw_out.txt","+w")
    swout_file.close()

    data_file = open("Ports.csv","w+", newline='')
    first_row = ["Hostname","Hostaddress","Port","PoE","Status","Description","VLAN"]
    writer = csv.DictWriter(data_file, fieldnames=first_row)
    writer.writeheader()

    archivo_excel = load_workbook("database.xlsx") 
    pestaña_excel = archivo_excel["Ports"]
    for x in range(2, 999999):
        valor = pestaña_excel.cell(row=x, column=1).value
        if valor != None and valor not in sw_ios:
            sw_ios.append(valor)
        elif valor == None:
            break

    tiempo1 = datetime.now()
    tiempo_inicial = tiempo1.strftime("%H:%M:%S")
    print(f"Hora de inicio: {tiempo_inicial}",f"Total de equipos a validar: {str(len(sw_ios))}",sep="\n")
    with concurrent.futures.ThreadPoolExecutor(max_workers=25) as executor:
        ejecucion = {executor.submit(connection,sw,user,pas,sw_out,swout_file,writer): sw for sw in sw_ios}
    for output in concurrent.futures.as_completed(ejecucion):
            output.result()
    data_file.close()

    date = tiempo1.strftime("%Y%m%d")
    csv_file = read_csv("Ports.csv")
    csv_file.to_excel(f"Ports {date}.xlsx",index=None,header=True,freeze_panes=(1,0))
    os.remove("Ports.csv")

    tiempo2 = datetime.now()
    tiempo_final = tiempo2.strftime("%H:%M:%S")
    tiempo_ejecucion = tiempo2 - tiempo1
    print(f"Hora de finalizacion: {tiempo_final}", f"Tiempo de ejecucion: {tiempo_ejecucion}", f"Total de equipos validados: {str(len(sw_ios))}",
    f"Total de equipos fuera: {str(len(sw_out))}",sep="\n")

if __name__ == "__main__":
    main()