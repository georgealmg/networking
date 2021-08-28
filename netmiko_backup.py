#!/usr/bin/env python3
#v1.0.2

import os, concurrent.futures, sys
from openpyxl import load_workbook
from getpass import getpass
from datetime import datetime
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoTimeoutException, SSHException, AuthenticationException

def backup(sw,user,pas,sw_out):
    try:
        conn = ConnectHandler(device_type= "cisco_ios_ssh",host= sw,username= user,password= pas)
    except(AuthenticationException):
        sys.exit("Por favor ejecute nuevamente el programa ya que introdujo una contraseña erronea.")
    except(NetMikoTimeoutException,SSHException):
        try:
            conn = ConnectHandler(device_type= "cisco_ios_telnet",host= sw,username= user,password= pas)
        except:
            print(f"Switch con IP \"{sw}\" esta fuera de linea.")
            sw_out.append(sw)
            return None
    hostname = conn.find_prompt()
    respaldos_txt = open(hostname.strip("#")+".txt","+w")
    respaldos_txt.write(conn.send_command("show running-config"))
    respaldos_txt.close()
    try:
        conn.save_config()
    except(OSError):
        conn.save_config(confirm=True,confirm_response="yes\r")
    conn.disconnect()
    print(f"{hostname} configuracion respaldada.")

def main():
    windowsuser = os.getlogin()
    try:
        os.chdir("C:/Users/"+str(windowsuser)+"/Documents")
    except(FileNotFoundError):
        os.chdir(os.getcwd())
    user = input("Username: ")
    pas = getpass()

    sw_dc = []
    sw_out = []
    excel_file = load_workbook("Respaldos.xlsx")
    sheet = excel_file["IP"]
    for ip in range(2, 999999):
        valor = sheet.cell(row=ip, column=1).value
        if valor != None and valor not in sw_dc:
            sw_dc.append(valor)
        elif valor == None:
            break
    tiempo1 = datetime.now()
    tiempo_inicial = tiempo1.strftime("%H:%M:%S")
    print(f"La ejecucion de este programa inicio a las {tiempo_inicial}, se respaldara la configuracion de {str(len(sw_dc))} equipos.")
    
    fecha = datetime.now()
    directorio_fecha = fecha.strftime("%d-%m-%y")
    directorio = "Data Center"
    try:
        os.mkdir(f"C:/Users/{windowsuser}/Documents/{directorio}")
    except(FileExistsError):
        print(f"La carpeta {directorio} ya existe.")
    try:
        os.mkdir(f"C:/Users/{windowsuser}/Documents/{directorio}/{directorio_fecha}")
    except(FileExistsError):
        print(f"La carpeta {directorio}/{directorio_fecha} ya existe.")
    os.chdir(f"C:/Users/{windowsuser}/Documents/{directorio}/{directorio_fecha}")
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor_nx:
        ejecucion_nx = {executor_nx.submit(backup,sw,user,pas,sw_out,directorio,directorio_fecha): sw for sw in sw_dc}
    for output_nx in concurrent.futures.as_completed(ejecucion_nx):
        output_nx.result()

    with open("sw_out.txt","+w") as file:
        for sw in sw_out:
            file.write(sw)

    tiempo2 = datetime.now()
    tiempo_final = tiempo2.strftime("%H:%M:%S")
    print(f"La ejecucion de este programa finalizo a las {tiempo_final}, se respaldo la configuracion de {str(len(sw_dc))} equipos, {str(len(sw_out))} de los equipos estan fuera de linea.")
    tiempo_ejecucion = tiempo2 - tiempo1
    print(f"El tiempo de ejecucion del programa fue de: {tiempo_ejecucion}")

if __name__ == "__main__":
    main()
