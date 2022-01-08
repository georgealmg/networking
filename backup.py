#!/usr/bin/env python3
#v1.0.4

import concurrent.futures, os, socket
from getpass import getpass, getuser
from datetime import datetime
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetmikoTimeoutException, SSHException, AuthenticationException
from openpyxl import load_workbook

try:
   os.chdir(f"/mnt/c/Users/{getuser()}/Documents")
except(FileNotFoundError):
    os.chdir(os.getcwd())

user = input("Username: ")
pas = getpass()
sw_ios = []
sw_out = []
swout_file = open("sw_out.txt","w")
swout_file.close()

excel_file = load_workbook("Backup.xlsx")
sheet = excel_file["Devices"]
for ip in range(2, 999999):
    valor = sheet.cell(row=ip, column=1).value
    if valor != None and valor not in sw_ios:
        sw_ios.append(valor)
    elif valor == None:
        break

def backup(conn,date):
    hostname = conn.find_prompt()
    try:
        respaldos_txt = open((f"/mnt/c/{getuser()}/Documents/Backup {date}/{hostname}+.txt","w"))
    except(FileNotFoundError):
        respaldos_txt = open((f"{os.getcwd()}/Documents/Backup {date}/{hostname}+.txt","w"))
    respaldos_txt.write(conn.send_command("show running-config"))
    respaldos_txt.close()
    try:
        conn.save_config()
    except(OSError):
        conn.save_config(confirm=True,confirm_response="yes\r")
    conn.disconnect()
    print(f"{hostname} configuracion respaldada.")

def connection(sw,date):
    try:
        conn = ConnectHandler(device_type= "cisco_ios_ssh",host= sw,username= user,password= pas, fast_cli= False)
        backup(conn,date)
    except(ConnectionRefusedError, ConnectionResetError):
        sw_out.append(sw)
        print(f"Error:{sw}:ConnectionRefused error")
        swout_file = open("sw_out.txt","a")
        swout_file.write(f"Error:{sw}:ConnectionRefused error"+"\n")
        swout_file.close()
    except(TimeoutError, socket.timetout):
        sw_out.append(sw)
        print(f"Error:{sw}:Timeout error")
        swout_file = open("sw_out.txt","a")
        swout_file.write(f"Error:{sw}:Timeout error"+"\n")
        swout_file.close()
    except(SSHException, NetmikoTimeoutException):
        try:
            conn = ConnectHandler(device_type= "cisco_ios_telnet",host= sw,username= user,password= pas,fast_cli= False)
            backup(conn,date)
        except(ConnectionRefusedError, ConnectionResetError):
            sw_out.append(sw)
            print(f"Error:{sw}:ConnectionRefused error")
            swout_file = open("sw_out.txt","a")
            swout_file.write(f"Error:{sw}:ConnectionRefused error"+"\n")
            swout_file.close()
        except(TimeoutError, socket.timetout):
            sw_out.append(sw)
            print(f"Error:{sw}:Timeout error")
            swout_file = open("sw_out.txt","a")
            swout_file.write(f"Error:{sw}:Timeout error"+"\n")
            swout_file.close()
        except(AuthenticationException):
            sw_out.append(sw)
            print(f"Error:{sw}:Authentication error")
            swout_file = open("sw_out.txt","a")
            swout_file.write(f"Error:{sw}:Authentication error"+"\n")
            swout_file.close()
    except(AuthenticationException):
        sw_out.append(sw)
        print(f"Error:{sw}:Authentication error")
        swout_file = open("sw_out.txt","a")
        swout_file.write(f"Error:{sw}:Authentication error"+"\n")
        swout_file.close()
    except(EOFError):
        sw_out.append(sw)
        print(f"Error:{sw}:EOF error")
        swout_file = open("sw_out.txt","a")
        swout_file.write(f"Error:{sw}:EOF error"+"\n")
        swout_file.close()

def main():

    tiempo1 = datetime.now()
    tiempo_inicial = tiempo1.strftime("%H:%M:%S")
    print(f"Hora de inicio: {tiempo_inicial}",f"Total de equipos a respaldar: {str(len(sw_ios))}",sep="\n")

    date = tiempo1.strftime("%d-%m-%y")
    try:
        os.mkdir(f"/mnt/c/{getuser()}/Documents/Backup {date}")
    except(FileExistsError):
        try:
            os.mkdir(f"{os.getcwd()}/Backup {date}")
        except(FileExistsError):
            pass
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        ejecucion = {executor.submit(backup,sw,date): sw for sw in sw_ios}
    for output in concurrent.futures.as_completed(ejecucion):
        output.result()

    tiempo2 = datetime.now()
    tiempo_final = tiempo2.strftime("%H:%M:%S")
    tiempo_ejecucion = tiempo2 - tiempo1
    print(f"Hora de finalizacion: {tiempo_final}", f"Tiempo de ejecucion: {tiempo_ejecucion}", f"Total de configuraciones respaldadas: {str(len(sw_ios)-len(sw_out))}",
    f"Total de equipos fuera: {str(len(sw_out))}",sep="\n")

if __name__ == "__main__":
    main()
