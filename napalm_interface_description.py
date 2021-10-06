#!/usr/bin/env python3
#v1.0.0

import concurrent.futures, csv, os, sys, time
from getpass import getpass, getuser
from datetime import datetime
from napalm import get_network_driver
from napalm.base.exceptions import ConnectionException
from netmiko.ssh_exception import AuthenticationException
from pandas import read_csv
from openpyxl import load_workbook


def inter_description(conn,sw,writer):
    hostname = conn.get_facts()["hostname"]
    print(f"Validacion iniciada --> {hostname}.")
    for inter in conn.get_interfaces():
        time.sleep(1)
        writer.writerow({"Hostname": hostname ,"IP":sw ,"Interface":inter ,"Description":conn.get_interfaces()[inter]["description"]})
    print(f"Validacion finalizada --> {hostname}.")
    conn.close()

def connection(sw,user,pas,sw_out,writer):
    driver = get_network_driver("ios")
    try:
        conn = driver(hostname= sw,username= user, password= pas)
        conn.open()
        inter_description(conn,sw,writer)
    except(ConnectionRefusedError):
        sw_out.append(sw)
        print(f"Error:{sw}:ConnectionRefused error")
        swout_file = open("sw_out.txt","a")
        swout_file.write(f"Error:{sw}:ConnectionRefused error"+"\n")
        swout_file.close()
    except(AuthenticationException):
        sw_out.append(sw)
        print(f"Error:{sw}:Authentication error")
        swout_file = open("sw_out.txt","a")
        swout_file.write(f"Error:{sw}:Authentication error"+"\n")
        swout_file.close()
    except(ConnectionException):
        try:
            conn = driver(hostname= sw,username= user, password= pas, optional_args= {"transport": 'telnet'})
            conn.open()
            inter_description(conn,sw,writer)
        except(ConnectionRefusedError):
            sw_out.append(sw)
            print(f"Error:{sw}:ConnectionRefused error")
            swout_file = open("sw_out.txt","a")
            swout_file.write(f"Error:{sw}:ConnectionRefused error"+"\n")
            swout_file.close()
        except(TimeoutError):
            sw_out.append(sw)
            print(f"Error:{sw}:Timeout error")
            swout_file = open("sw_out.txt","a")
            swout_file.write(f"Error:{sw}:Timeout error"+"\n")
            swout_file.close()
        except(AuthenticationException):
            sw_out.append(sw)
            print(f"Error:{sw}:Authentication error")
            swout_file = open("sw_out.txt","a")
            swout_file.write(f"Error:{sw}:Authentication error"+"\n")
            swout_file.close()
    except(EOFError):
        sw_out.append(sw)
        print(f"Error:{sw}:EOF error")
        swout_file = open("sw_out.txt","a")
        swout_file.write(f"Error:{sw}:EOF error"+"\n")
        swout_file.close()

def main():
    try:
       os.chdir(f"/mnt/c/Users/{getuser()}/Documents/networking")
    except(FileNotFoundError):
        os.chdir(os.getcwd())
    user = input("Username: ")
    pas = getpass()
    sw_ios = []
    sw_out = []
    swout_file = open("sw_out.txt","w")
    swout_file.close()

    data_file = open("interfaces_descriptions.csv","w", newline='')
    first_row = ["Hostname","IP","Interface","Description"]
    writer = csv.DictWriter(data_file, fieldnames=first_row,dialect="excel")
    writer.writeheader()

    archivo_excel = load_workbook("database.xlsx") 
    pestaña_excel = archivo_excel["Ports"]
    for x in range(2, 999999):
        valor = pestaña_excel.cell(row=x, column=1).value
        if valor != None and valor not in sw_ios:
            sw_ios.append(valor)
        elif valor == None:
            break
    
    total_sw = len(sw_ios)
    tiempo1 = datetime.now()
    tiempo_inicial = tiempo1.strftime("%H:%M:%S")
    print(f"Hora de inicio: {tiempo_inicial}",f"Total de equipos a validar: {str(total_sw)}",sep="\n")
    with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor_ios:
        ejecucion_ios = {executor_ios.submit(connection,sw,user,pas,sw_out,writer): sw for sw in sw_ios}
    for output_ios in concurrent.futures.as_completed(ejecucion_ios):
        output_ios.result()
    data_file.close()

    csv_file = read_csv("interfaces_descriptions.csv")
    csv_file.to_excel(f"interfaces_descriptions.xlsx",index=None,header=True,freeze_panes=(1,0))
    os.remove("interfaces_descriptions.csv")

    contador_out = len(sw_out)
    tiempo2 = datetime.now()
    tiempo_final = tiempo2.strftime("%H:%M:%S")
    tiempo_ejecucion = tiempo2 - tiempo1
    print(f"Hora de finalizacion: {tiempo_final}", f"Tiempo de ejecucion: {tiempo_ejecucion}", f"Total de equipos validados: {str(total_sw)}",f"Total de equipos fuera: {str(contador_out)}",sep="\n")

if __name__ == "__main__":
    main()